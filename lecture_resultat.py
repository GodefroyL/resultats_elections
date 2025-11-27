from openpyxl import load_workbook


def lecture_resultat(fichier: str) -> dict:
    """
    Lecture des résultats des élections législatives à partir d'un fichier Excel.
    
    :param fichier: Fichier Excel contenant les résultats des élections législatives (ministère de l'intérieur)
    :return: Dictionnaire contenant les informations des circonscriptions et des candidats.
    """
# Charger le fichier Excel
    wb = load_workbook(filename=fichier, data_only=True)

# Sélectionner la feuille active (ou spécifier le nom de la feuille)
    feuille = wb.active

    resultats = {}
# Parcourir chaque ligne
    for ligne in feuille.iter_rows(values_only=True):
# Ignorer la ligne d'en-tête
        if ligne[0] == 'Code département':
            continue
    # Extraire les informations de la circonscription
        infos_circo = {
            'code_departement': ligne[0],
            'departement': ligne[1],
            'code_circonscription': ligne[2],
            'nom_circonscription': ligne[3],
            'nombres_inscrits': ligne[4],
            'nombres_votants': ligne[5],
            'pourcentage_votants': ligne[6],
            'abstentions': ligne[7],
            'pourcentage_abstentions': ligne[8],
            'exprimés': ligne[9],
            'pourcentage_exprimés/inscrit': ligne[10],
            'blacs_nuls': int(ligne[12])+float(ligne[15]),
            'infos_candidats': {}
        }
        
        i=18
    # Extraire les informations des candidats
        while i < len(ligne) and ligne[i] != None:
            infos_candidat = {
                'parti': ligne[i+1],
                'nom': f'{ligne[i+2]}, {ligne[i+3]}',
                'voix': ligne[i+5],
                'pourcentage/inscrit': ligne[i+6],
                'pourcentage/votant': ligne[i+7]
                }
            infos_circo['infos_candidats'][ligne[i]] = infos_candidat
            i += 9

        resultats[ligne[2]] = infos_circo
    return resultats


def get_resultats_circo(resultats: dict, circonscription_code: int, affichage_nom=True, affichage_parti=True)-> dict:
    """
    Récupère les résultats d'une circonscription spécifique.
    
    :param resultats: Dictionnaire des résultats des élections législatives.
    :param circonscription_code: Code de la circonscription à rechercher.
    :param nom: controle l'affichage des noms des candidats.
    :param parti: controle l'affichage des partis des candidats.
    :return: {id_candidat : {'parti': , 'nom': , 'voix': , 'pourcentage/inscrit': , 'pourcentage/votant': }}
    """
    infos_circo = resultats.get(circonscription_code)
    if infos_circo:
        for candidat_key, candidat_info in infos_circo['infos_candidats'].items():
            pourcentage_voix = candidat_info['pourcentage/votant']
            if affichage_nom:
                nom = candidat_info['nom']
                nom = f' ({nom})'
            if affichage_parti:
                parti = candidat_info['parti']
                parti = f'{parti}'
            if affichage_parti or affichage_nom:
                print(f'{parti}{nom} \t| {pourcentage_voix}')
                print('------------------------------------')
        return infos_circo.get('infos_candidats')


def get_resultats_parti(resultats: dict, code_circo: int, parti: str) -> tuple[str, bool, bool]:
    """
    Récupère les résultats d'un parti spécifique dans une circonscription donnée.
    
    :param resultats: Dictionnaire des résultats des élections législatives.
    :param code_circo: Code de la circonscription à rechercher.
    :param parti: Nom du parti politique à rechercher.
    :return: pourcentage_voix, second_tour_possible (bool), elu (bool)
    """
    resultat_circo = get_resultats_circo(resultats, code_circo, affichage_nom=False, affichage_parti=False)
    for candidat_info in resultat_circo.values():
        if candidat_info.get('parti') == parti:
            second_tour = False
            elu = False

            if float(candidat_info.get('pourcentage/inscrit').replace('%', '').replace(',', '.')) >= 25: elu = True
            if float(candidat_info.get('pourcentage/votant').replace('%', '').replace(',', '.')) >= 12.5: second_tour = True
            return candidat_info.get('pourcentage/votant'), second_tour, elu
    return '0%', False, False


def get_departements(resultats: dict) -> list:
    """
    Récupère la liste des départements présents dans le fichier des résultats.
    
    :param resultats: Dictionnaire des résultats des élections législatives.
    :return: Liste des départements.
    """
    set_departements = set()
    departements = []
    for circo_info in resultats.values():
        departements.append((circo_info.get('code_departement'), circo_info.get('departement')))
        set_departements.add((circo_info.get('code_departement'), circo_info.get('departement')))

    returned_departements = []
    for dep in departements:
        if dep in set_departements:
            returned_departements.append(dep)
            set_departements.remove(dep)

    return returned_departements


def get_circonscriptions(resultats: dict, code_departement: int) -> list:
    """
    Récupère la liste des circonscriptions pour un département donné.
    
    :param resultats: Dictionnaire des résultats des élections législatives.
    :param code_departement: Code du département à rechercher.
    :return: Liste des circonscriptions dans le département spécifié.
    """
    circonscriptions = []
    for circo_info in resultats.values():
        if circo_info.get('code_departement') == code_departement:
            circonscriptions.append((circo_info.get('code_circonscription'), circo_info.get('nom_circonscription')))
    return circonscriptions


def get_partis(resultats: dict) -> list:
    """
    Récupère la liste des partis politiques présents dans le fichier des résultats.
    
    :param resultats: Dictionnaire des résultats des élections législatives.
    :return: Liste des partis politiques.
    """
    partis = set()
    for circo_info in resultats.values():
        for candidat_info in circo_info['infos_candidats'].values():
            partis.add(candidat_info['parti'])
    return list(partis)


def get_partis_circonscription(resultats: dict, code_circo: int) -> list:
    """
    Récupère la liste des partis politiques présents dans une circonscription donnée.
    
    :param resultats: Dictionnaire des résultats des élections législatives.
    :param code_circo: Code de la circonscription à rechercher.
    :return: Liste des partis politiques dans la circonscription spécifiée.
    """
    resultat_circo = get_resultats_circo(resultats, code_circo, affichage_nom=False, affichage_parti=False)
    partis = set()
    for candidat_info in resultat_circo.values():
        partis.add(candidat_info['parti'])
    return list(partis)


def elu_premier_tour(resultats: dict, code_circo: int) -> tuple[bool, str, str]:
    """
    Vérifie si un candidat est élu au premier tour dans une circonscription donnée.
    
    :param resultats: Dictionnaire des résultats des élections législatives.
    :param code_circo: Code de la circonscription à rechercher.
    :return: True, parti élu, score si un candidat est élu au premier tour, False, '', '0%' sinon.
    """
    resultat_circo = get_resultats_circo(resultats, code_circo, affichage_nom=False, affichage_parti=False)
    for candidat_info in resultat_circo.values():
        elu = False
        parti_elu = ''
        score = '0%'
        if float(candidat_info.get('pourcentage/inscrit').replace('%', '').replace(',', '.')) >= 25 and float(candidat_info.get('pourcentage/votant').replace('%', '').replace(',', '.')) >= 50:
            elu = True
            parti_elu = candidat_info.get('parti')
            score = candidat_info.get('pourcentage/votant')
            break
    return elu, parti_elu, score


if __name__ == "__main__":
    fichier = 'C:/Users/godef/Downloads/resultats-definitifs-par-circonscriptions-legislatives.xlsx'
    resultat = lecture_resultat(fichier)
    print(get_resultats_circo(resultat, 202))